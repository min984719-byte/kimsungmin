#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
generate.py — SYNOPEX 설비 가동율 대시보드 자동 생성
template.html 이 없어도 기존 설비가동율대시보드.html 을 자동 갱신
"""
import json, os, glob, re
from openpyxl import load_workbook

print("="*55)
print("  SYNOPEX 설비 가동율 대시보드 자동 생성")
print("="*55)

OUTPUT = '설비가동율대시보드.html'

# ── 엑셀 파일 찾기 ──
files = sorted(glob.glob('data/*.xlsx') + glob.glob('data/*.xls'))
print(f"엑셀 목록: {files}")
if not files:
    print("❌ data/ 폴더에 엑셀 파일 없음"); exit(1)
excel_path = files[-1]
print(f"✅ 사용 파일: {excel_path}")

# ── 파싱 ──
def sf(v):
    try: return float(v) if v is not None else 0
    except: return 0

STOP = ['FPCB설비','합계','합 계','계획비가동','Trouble','비가동현황']

def parse_sheet(ws):
    rows = list(ws.iter_rows(min_row=1, max_row=80, values_only=True))
    period=''; overall=0; summary=[]; detail=[]; started=False
    for row in rows:
        c2 = str(row[2]).strip().replace('\n',' ') if row[2] is not None else ''
        c3 = str(row[3]).strip() if row[3] is not None else ''
        if not period and c2 and c2[0].isdigit() and not started:
            period=c2; rr=sf(row[9])
            overall=round(rr*100,1) if rr<=1 else round(rr,1); continue
        if c2=='공장명': started=True; continue
        if not started or not c2: continue
        proc=c2.replace(' ','')
        if any(s in proc for s in STOP): break
        sub=c3 if c3 else None
        rr2=sf(row[8]); rate=round(rr2*100,1) if rr2<=1 else round(rr2,1)
        item={'proc':proc,'sub':sub,'cnt':int(sf(row[4])) if row[4] else 0,
              'area':round(sf(row[5]),1),'base':round(sf(row[6])),'run':round(sf(row[7])),
              'rate':rate,'plan_down':round(sf(row[9])),'op_loss':round(sf(row[10])),
              'eq_loss':round(sf(row[11])),'eq_fail':round(sf(row[12]))}
        if not sub: summary.append(item)
        else:       detail.append(item)
    if overall==0 and summary:
        vv=[s['rate'] for s in summary if s['rate']>0]
        overall=round(sum(vv)/len(vv),1) if vv else 0
    return {'summary':summary,'detail':detail,'period':period,'overall':overall}

wb = load_workbook(excel_path, read_only=True, data_only=True)
print(f"시트: {wb.sheetnames}")

sheet_m  = next((s for s in wb.sheetnames if '월 누적' in s), None) or \
           next((s for s in wb.sheetnames if '누적' in s and '(일)' not in s), None)
sheet_d  = next((s for s in wb.sheetnames if '(일)' in s), None)
sheet_yr = next((s for s in wb.sheetnames if '2026' in s or '2027' in s), None)

m_data = parse_sheet(wb[sheet_m]) if sheet_m else {'summary':[],'detail':[],'period':'','overall':0}
d_data = parse_sheet(wb[sheet_d]) if sheet_d else {'summary':[],'detail':[],'period':'','overall':0}
print(f"✅ 월 누적: {m_data['period']} / {m_data['overall']}% / {len(m_data['detail'])}설비")
print(f"✅ 일 기준: {d_data['period']} / {d_data['overall']}% / {len(d_data['detail'])}설비")

trend={}; etrd={}
if sheet_yr:
    rows_yr=list(wb[sheet_yr].iter_rows(min_row=1,max_row=80,values_only=True))
    started=False
    for row in rows_yr:
        c2=str(row[2]).strip().replace('\n','').replace(' ','') if row[2] else ''
        if c2=='공장명': started=True; continue
        if not started or not c2: continue
        if any(s in c2 for s in STOP): break
        sub=str(row[3]).strip() if row[3] else None
        vals=[]
        for mi in range(4):
            v=sf(row[6+mi])
            vals.append(round(v*100,1) if 0<v<=1 else (round(v,1) if v>0 else None))
        if not any(v is not None for v in vals): continue
        if not sub: trend[c2]=vals
        else:       etrd[c2+'_'+sub]=vals
wb.close()

# ── 새 데이터 JS 블록 ──
new_data_block = (
    "<script>\n// ── 데이터 ──\n"
    "var MODES = " + json.dumps({'month':m_data,'day':d_data}, ensure_ascii=False) + ";\n"
    "var TREND = " + json.dumps(trend, ensure_ascii=False) + ";\n"
    "var ETRD  = " + json.dumps(etrd,  ensure_ascii=False) + ";"
)

# ── HTML 파일 처리 ──
if not os.path.exists(OUTPUT):
    # template.html 로 시도
    if os.path.exists('template.html'):
        with open('template.html', encoding='utf-8') as f:
            html = f.read()
        if '/* DATA_PLACEHOLDER */' in html:
            html = html.replace(
                '<script>\n// ── 데이터 (자동 생성) ──\n/* DATA_PLACEHOLDER */',
                new_data_block
            )
            print("✅ template.html 마커 방식 사용")
        else:
            print("❌ template.html 에 마커 없음"); exit(1)
    else:
        print("❌ 설비가동율대시보드.html 와 template.html 모두 없음"); exit(1)
else:
    with open(OUTPUT, encoding='utf-8') as f:
        html = f.read()
    
    # 방법1: 마커 방식
    if '/* DATA_PLACEHOLDER */' in html:
        html = html.replace(
            '<script>\n// ── 데이터 (자동 생성) ──\n/* DATA_PLACEHOLDER */',
            new_data_block
        )
        print("✅ 마커 방식으로 데이터 교체")
    
    # 방법2: 기존 데이터 블록 교체 (regex)
    else:
        pattern = re.compile(
            r'<script>\n// ── 데이터 ──\nvar MODES = \{.*?\};\nvar TREND = \{.*?\};\nvar ETRD\s*=\s*\{.*?\};',
            re.DOTALL
        )
        m = pattern.search(html)
        if m:
            html = html[:m.start()] + new_data_block + html[m.end():]
            print("✅ 정규식 방식으로 데이터 교체")
        else:
            # 방법3: MODES 변수만 교체
            m2 = re.search(r'var MODES = \{.*?\};', html, re.DOTALL)
            m3 = re.search(r'var TREND = \{.*?\};', html, re.DOTALL)
            m4 = re.search(r'var ETRD\s*=\s*\{.*?\};', html, re.DOTALL)
            if m2 and m3 and m4:
                html = html[:m2.start()] + \
                    "var MODES = " + json.dumps({'month':m_data,'day':d_data}, ensure_ascii=False) + ";\n" + \
                    html[m2.end():m3.start()] + \
                    "var TREND = " + json.dumps(trend, ensure_ascii=False) + ";\n" + \
                    html[m3.end():m4.start()] + \
                    "var ETRD  = " + json.dumps(etrd, ensure_ascii=False) + ";" + \
                    html[m4.end():]
                print("✅ 변수별 개별 교체 방식 사용")
            else:
                print(f"❌ 데이터 교체 실패. MODES:{bool(m2)} TREND:{bool(m3)} ETRD:{bool(m4)}")
                exit(1)

with open(OUTPUT, 'w', encoding='utf-8') as f:
    f.write(html)
print(f"✅ 생성 완료: {OUTPUT} ({len(html)//1024}KB)")
