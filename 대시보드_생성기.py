#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
SYNOPEX 설비 가동율 대시보드 생성기
====================================
이 파일을 더블클릭하면 자동으로 대시보드 HTML이 생성됩니다.

사용법:
  1. 이 파일과 같은 폴더에 엑셀 파일을 넣으세요
  2. 이 파일을 더블클릭하세요
  3. 생성된 HTML 파일을 GitHub에 업로드하세요
"""

import json, os, glob, sys, webbrowser
from pathlib import Path

# ── 실행 폴더 기준으로 경로 설정 ──
BASE_DIR = Path(__file__).parent

print("=" * 50)
print("  SYNOPEX 설비 가동율 대시보드 생성기")
print("=" * 50)

# ── 엑셀 파일 찾기 ──
excel_files = sorted(
    list(BASE_DIR.glob('*.xlsx')) +
    list(BASE_DIR.glob('*.xls')) +
    list(BASE_DIR.glob('data/*.xlsx')) +
    list(BASE_DIR.glob('data/*.xls'))
)

if not excel_files:
    print("\n❌ 엑셀 파일을 찾을 수 없습니다.")
    print("   이 파일과 같은 폴더에 엑셀 파일을 넣어주세요.")
    input("\n아무 키나 누르면 종료됩니다...")
    sys.exit(1)

# 여러 파일이 있으면 선택
if len(excel_files) == 1:
    excel_path = excel_files[0]
    print(f"\n✅ 파일 발견: {excel_path.name}")
else:
    print("\n📂 엑셀 파일 목록:")
    for i, f in enumerate(excel_files):
        print(f"   {i+1}. {f.name}")
    while True:
        try:
            choice = int(input("\n사용할 파일 번호를 입력하세요: ")) - 1
            if 0 <= choice < len(excel_files):
                excel_path = excel_files[choice]
                break
        except: pass
        print("올바른 번호를 입력하세요.")

print(f"\n📊 파일 읽는 중: {excel_path.name}")

# ── openpyxl 설치 확인 ──
try:
    from openpyxl import load_workbook
except ImportError:
    print("\n⚙️  필요한 패키지 설치 중...")
    import subprocess
    subprocess.run([sys.executable, '-m', 'pip', 'install', 'openpyxl'], check=True)
    from openpyxl import load_workbook
    print("✅ 설치 완료")

# ── 데이터 파싱 ──
def sf(v):
    try: return float(v) if v is not None else 0
    except: return 0

STOP = ['FPCB설비','합계','합 계','계획비가동','Trouble','비가동현황']

def parse_sheet(ws):
    rows = list(ws.iter_rows(min_row=1, max_row=80, values_only=True))
    period=''; overall=0; summary=[]; detail=[]; started=False
    for row in rows:
        c2 = str(row[2]).strip().replace('\n',' ') if row[2] is not None else ''
        c3 = str(row[3]).strip() if row[3] is not None else ''
        if not period and c2 and c2[0].isdigit() and not started:
            period=c2; rr=sf(row[9])
            overall=round(rr*100,1) if rr<=1 else round(rr,1); continue
        if c2=='공장명': started=True; continue
        if not started or not c2: continue
        proc=c2.replace(' ','')
        if any(s in proc for s in STOP): break
        sub=c3 if c3 else None
        rr2=sf(row[8]); rate=round(rr2*100,1) if rr2<=1 else round(rr2,1)
        item={'proc':proc,'sub':sub,'cnt':int(sf(row[4])) if row[4] else 0,
              'area':round(sf(row[5]),1),'base':round(sf(row[6])),'run':round(sf(row[7])),
              'rate':rate,'plan_down':round(sf(row[9])),'op_loss':round(sf(row[10])),
              'eq_loss':round(sf(row[11])),'eq_fail':round(sf(row[12]))}
        if not sub: summary.append(item)
        else:       detail.append(item)
    if overall==0 and summary:
        vv=[s['rate'] for s in summary if s['rate']>0]
        overall=round(sum(vv)/len(vv),1) if vv else 0
    return {'summary':summary,'detail':detail,'period':period,'overall':overall}

wb = load_workbook(str(excel_path), read_only=True, data_only=True)

sheet_m  = next((s for s in wb.sheetnames if '월 누적' in s), None) or \
           next((s for s in wb.sheetnames if '누적' in s and '(일)' not in s), None)
sheet_d  = next((s for s in wb.sheetnames if '(일)' in s), None)
sheet_yr = next((s for s in wb.sheetnames if '2026' in s or '2027' in s), None)

m_data = parse_sheet(wb[sheet_m]) if sheet_m else {'summary':[],'detail':[],'period':'','overall':0}
d_data = parse_sheet(wb[sheet_d]) if sheet_d else {'summary':[],'detail':[],'period':'','overall':0}

trend_m={}; equip_trend={}
if sheet_yr:
    rows_yr=list(wb[sheet_yr].iter_rows(min_row=1,max_row=80,values_only=True))
    started=False
    for row in rows_yr:
        c2=str(row[2]).strip().replace('\n','').replace(' ','') if row[2] else ''
        if c2=='공장명': started=True; continue
        if not started or not c2: continue
        if any(s in c2 for s in STOP): break
        sub=str(row[3]).strip() if row[3] else None
        vals=[]
        for mi in range(4):
            v=sf(row[6+mi])
            vals.append(round(v*100,1) if 0<v<=1 else (round(v,1) if v>0 else None))
        if not any(v is not None for v in vals): continue
        if not sub: trend_m[c2]=vals
        else: equip_trend[c2+'_'+sub]=vals

wb.close()

DATA={'month':m_data,'day':d_data,'trend':trend_m,'equip_trend':equip_trend,
      'source_file':excel_path.name}

print(f"✅ 월 누적: {m_data['period']} / {m_data['overall']}% / {len(m_data['detail'])}개 설비")
print(f"✅ 일 기준: {d_data['period']} / {d_data['overall']}% / {len(d_data['detail'])}개 설비")

# ── template.html 찾기 ──
template_path = BASE_DIR / 'template.html'
if not template_path.exists():
    print(f"\n❌ template.html 파일이 없습니다.")
    print(f"   이 파일과 같은 폴더에 template.html 을 넣어주세요.")
    input("\n아무 키나 누르면 종료됩니다...")
    sys.exit(1)

with open(template_path, encoding='utf-8') as f:
    html = f.read()

# ── 데이터 주입 ──
data_js = ("<script>\n/* ── 자동생성 " + m_data['period'] + " ── */\n"
           "var EMBEDDED=" + json.dumps(DATA, ensure_ascii=False) + ";\n</script>\n")
html = html.replace('</head>', data_js + '</head>', 1)

# 자동복원 로직 교체
MARKER="// ════════════════════════════════\n//  localStorage 자동 복원\n// ════════════════════════════════\n(function(){"
start=html.find(MARKER)
end=html.find('})();', start)+5 if start>0 else -1

NEW="""// ═══════════════════════
//  초기 로드: 내장 데이터 우선
// ═══════════════════════
(function(){
  if(typeof EMBEDDED!=='undefined'&&EMBEDDED&&EMBEDDED.month&&
     EMBEDDED.month.summary&&EMBEDDED.month.summary.length){
    DB.month=EMBEDDED.month; DB.day=EMBEDDED.day;
    TREND={month:EMBEDDED.trend,day:EMBEDDED.trend};
    EQUIP_TREND=EMBEDDED.equip_trend||{};
    var tog=document.getElementById('modeToggle');
    if(tog){tog.style.opacity='1';tog.style.pointerEvents='auto';}
    showSavedRow(EMBEDDED.source_file,EMBEDDED.month.period+' 기준');
    setStatus('✅ '+EMBEDDED.month.period+' 설비 가동율 자동 로드','ok');
    MODE='month';
    document.getElementById('btnMonth').classList.add('on');
    document.getElementById('btnDay').classList.remove('on');
    renderAll(); return;
  }
  try{
    var m=localStorage.getItem(LS_M); if(!m) return;
    DB.month=JSON.parse(m);
    var d=localStorage.getItem(LS_D); if(d) DB.day=JSON.parse(d);
    var et=localStorage.getItem(LS_ET); if(et) EQUIP_TREND=JSON.parse(et);
    var tr=localStorage.getItem(LS_TR); if(tr) TREND=JSON.parse(tr);
    if(!DB.month||!DB.month.summary||!DB.month.summary.length) return;
    var tog2=document.getElementById('modeToggle');
    if(tog2){tog2.style.opacity='1';tog2.style.pointerEvents='auto';}
    showSavedRow(localStorage.getItem('syn_oee_fname')||'',
                 localStorage.getItem('syn_oee_saved')||'');
    setStatus('💾 저장된 데이터 복원','ok');
    MODE='month';
    document.getElementById('btnMonth').classList.add('on');
    document.getElementById('btnDay').classList.remove('on');
    renderAll();
  }catch(e){console.warn('복원 실패:',e);}
})();"""

if start>0 and end>0:
    html=html[:start]+NEW+html[end:]

# ── 출력 ──
output_name = f"설비가동율대시보드.html"
output_path = BASE_DIR / output_name

with open(output_path, 'w', encoding='utf-8') as f:
    f.write(html)

print(f"\n{'='*50}")
print(f"🎉 완료! → {output_name}")
print(f"{'='*50}")
print(f"\n다음 단계:")
print(f"  1. {output_name} 파일을 GitHub에 업로드")
print(f"     github.com/min984719-byte/kimsungmin")
print(f"     → Add file → Upload files")
print(f"  2. Commit changes 클릭")
print(f"  3. 완료! 누구나 링크로 바로 열람 가능")

# 생성된 파일 자동으로 브라우저에서 열기
open_browser = input("\n생성된 파일을 브라우저로 바로 확인할까요? (y/n): ").strip().lower()
if open_browser == 'y':
    webbrowser.open(output_path.as_uri())

input("\n아무 키나 누르면 종료됩니다...")
